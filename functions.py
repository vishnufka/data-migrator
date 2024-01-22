import config as cfg
import pandas as pd
import sys
from openpyxl.worksheet.table import Table
from openpyxl.worksheet.table import TableStyleInfo
from openpyxl.utils import get_column_letter


def read_excel(filename, header_row_index, sheet_name=0):
    df = pd.read_excel(filename, header=header_row_index, sheet_name=sheet_name)
    return df


def prepare_df(df, filename):
    mapping_dict = {}
    columns_to_drop = []
    for key, value in cfg.mapping_dict[cfg.data_mapping_dict[filename]].items():
        if value:
            mapping_dict[key] = value
        else:
            columns_to_drop.append(key)

    df.rename(columns=mapping_dict, inplace=True)
    df.drop(columns=[c for c in columns_to_drop if c in df.columns], axis=1, inplace=True)

    return df


def link_x_to_y(df, linking_file_path, linking_pair, matching_column=None, not_hex=False):
    linking_df = read_excel(linking_file_path, cfg.header_row_index)
    linking_df = get_linking_df_as_pair(linking_df, linking_pair, not_hex)
    df = vlookup(df, linking_df, matching_column)
    return df


def get_linking_df_as_pair(linking_df, linking_pair, not_hex):
    linking_df = linking_df[linking_pair]
    linking_df = linking_df.fillna("0")
    if not_hex:
        pass
    else:
        try:
            linking_df[linking_pair[0]].apply(lambda x: int(x, 16))
        except ValueError:
            print("ABORT: Expecting hexadecimal")
            sys.exit()
    linking_df.set_index(linking_pair[0], drop=True, inplace=True)
    return linking_df


def vlookup(df, linking_df, matching_column):
    if matching_column:
        df = df.merge(linking_df, left_on=matching_column, right_index=True, how='left')
    else:
        df = df.merge(linking_df, left_index=True, right_index=True, how='left')
    return df


def export_to_excel(df, filename):
    if not isinstance(df, dict):
        df = {cfg.data_mapping_dict[filename]: df}
    with pd.ExcelWriter(cfg.output_path +
                        cfg.today_timestamp +
                        cfg.data_mapping_dict[filename] + ".xlsx",
                        engine="openpyxl") as writer:
        for sheet_name in df:
            df[sheet_name].to_excel(excel_writer=writer, sheet_name=sheet_name, index=False)
            sheet = writer.sheets[sheet_name]
            tab = Table(displayName=sheet_name.replace(" ", "_"),
                        ref="A1:" + get_column_letter(sheet.max_column) + str(sheet.max_row),
                        tableStyleInfo=TableStyleInfo(name='TableStyleMedium2', showRowStripes=True))
            sheet.add_table(tab)


def send_df_to_client_method(df, filename):
    package = cfg.client_folder_name + "." + cfg.data_mapping_dict[filename]
    name = cfg.data_mapping_dict[filename]
    client_method = getattr(_import_(package, fromlist=[name]), name)
    df = client_method(df)
    return df


def run_extra_client_methods():
    package = cfg.client_folder_name + ".extra_client_methods"
    name = "extra_client_methods"
    df = None
    try:
        client_method = getattr(_import_(package, fromlist=[name]), name)
        df = client_method()
    except ModuleNotFoundError:
        pass

    return df
